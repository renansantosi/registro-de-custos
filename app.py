import os
import uuid
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, send_file, redirect, url_for, render_template
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
import threading
from database import (
    init_db, get_imoveis, add_imovel, get_custos, get_all_custos,
    add_custo, update_custo, delete_custo, get_custo,
    add_usuario, get_usuario_by_email, get_usuario_by_id,
    change_password, create_reset_token, get_valid_reset_token, delete_reset_token,
    update_usuario_avatar,
    get_imovel_by_id, get_parceiros, add_parceiro, delete_parceiro, update_parceiro_notificar,
    save_viabilidade, get_viabilidades, get_viabilidade_by_id, delete_viabilidade
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
DB_PATH = os.path.join(BASE_DIR, 'registro.db')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-leialopro-change-in-prod')
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

login_manager = LoginManager(app)
login_manager.login_view = 'login_page'

app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', '')
mail = Mail(app)

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}


# ─── User model ──────────────────────────────────────────────────────────────

class User(UserMixin):
    def __init__(self, row):
        self.id = row['id']
        self.nome = row['nome']
        self.email = row['email']


@login_manager.user_loader
def load_user(user_id):
    row = get_usuario_by_id(DB_PATH, user_id)
    return User(row) if row else None


@login_manager.unauthorized_handler
def unauthorized():
    if request.path.startswith('/api/') or request.is_json:
        return jsonify({'error': 'Sessao expirada. Faca login novamente.'}), 401
    return redirect(url_for('login_page'))


# ─── Helpers ─────────────────────────────────────────────────────────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_upload(file):
    if not file or file.filename == '':
        return None, None, None
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    unique_name = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex
    save_path = os.path.join(UPLOAD_FOLDER, unique_name)
    file.save(save_path)
    return unique_name, file.filename, file.content_type


def imovel_belongs_to_user(imovel_id, user_id):
    imoveis = get_imoveis(DB_PATH, user_id)
    return any(im['id'] == imovel_id for im in imoveis)


def custo_belongs_to_user(custo_id, user_id):
    custo = get_custo(DB_PATH, custo_id)
    if not custo:
        return False
    return imovel_belongs_to_user(custo['imovel_id'], user_id)


# ─── Auth pages ──────────────────────────────────────────────────────────────

@app.route('/login')
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json(force=True) or {}
    email = (data.get('email') or '').strip().lower()
    senha = data.get('senha') or ''
    if not email or not senha:
        return jsonify({'error': 'E-mail e senha sao obrigatorios'}), 400
    usuario = get_usuario_by_email(DB_PATH, email)
    if not usuario or not check_password_hash(usuario['senha_hash'], senha):
        return jsonify({'error': 'E-mail ou senha incorretos'}), 401
    user = User(usuario)
    login_user(user, remember=True)
    return jsonify({'ok': True, 'nome': user.nome})


@app.route('/api/auth/register', methods=['POST'])
def api_register():
    data = request.get_json(force=True) or {}
    nome = (data.get('nome') or '').strip()
    email = (data.get('email') or '').strip().lower()
    senha = data.get('senha') or ''
    if not nome or not email or not senha:
        return jsonify({'error': 'Nome, e-mail e senha sao obrigatorios'}), 400
    if len(senha) < 6:
        return jsonify({'error': 'Senha deve ter pelo menos 6 caracteres'}), 400
    if get_usuario_by_email(DB_PATH, email):
        return jsonify({'error': 'E-mail ja cadastrado'}), 409
    senha_hash = generate_password_hash(senha)
    usuario = add_usuario(DB_PATH, nome, email, senha_hash)
    user = User(usuario)
    login_user(user, remember=True)
    return jsonify({'ok': True, 'nome': user.nome}), 201


@app.route('/api/auth/logout', methods=['POST'])
@login_required
def api_logout():
    logout_user()
    return jsonify({'ok': True})


@app.route('/api/auth/me')
@login_required
def api_me():
    usuario = get_usuario_by_id(DB_PATH, current_user.id)
    avatar_url = None
    if usuario and usuario.get('avatar_path'):
        avatar_url = url_for('serve_upload', filename=usuario['avatar_path'])
    return jsonify({'id': current_user.id, 'nome': current_user.nome, 'email': current_user.email, 'avatar_url': avatar_url})


@app.route('/api/auth/upload-avatar', methods=['POST'])
@login_required
def api_upload_avatar():
    file = request.files.get('avatar')
    if not file or file.filename == '':
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'Formato nao permitido. Use JPG, PNG ou WEBP.'}), 400
    usuario = get_usuario_by_id(DB_PATH, current_user.id)
    if usuario and usuario.get('avatar_path'):
        old = os.path.join(UPLOAD_FOLDER, usuario['avatar_path'])
        if os.path.exists(old):
            os.remove(old)
    unique_name, _, _ = save_upload(file)
    update_usuario_avatar(DB_PATH, current_user.id, unique_name)
    avatar_url = url_for('serve_upload', filename=unique_name)
    return jsonify({'ok': True, 'avatar_url': avatar_url})


@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.get_json(force=True) or {}
    senha_atual = data.get('senha_atual', '')
    senha_nova = data.get('senha_nova', '')
    if not senha_atual or not senha_nova:
        return jsonify({'error': 'Preencha todos os campos'}), 400
    if len(senha_nova) < 6:
        return jsonify({'error': 'Nova senha deve ter pelo menos 6 caracteres'}), 400
    usuario = get_usuario_by_id(DB_PATH, current_user.id)
    if not check_password_hash(usuario['senha_hash'], senha_atual):
        return jsonify({'error': 'Senha atual incorreta'}), 401
    change_password(DB_PATH, current_user.id, generate_password_hash(senha_nova))
    return jsonify({'ok': True})


@app.route('/api/auth/forgot-password', methods=['POST'])
def api_forgot_password():
    data = request.get_json(force=True) or {}
    email = (data.get('email') or '').strip().lower()
    if not email:
        return jsonify({'error': 'Informe o e-mail'}), 400
    usuario = get_usuario_by_email(DB_PATH, email)
    if not usuario:
        return jsonify({'ok': True})  # nao revela se email existe
    if not app.config.get('MAIL_USERNAME'):
        return jsonify({'error': 'Envio de email nao configurado. Contate o administrador.'}), 503
    token = create_reset_token(DB_PATH, usuario['id'])
    base_url = request.host_url.rstrip('/')
    reset_url = f"{base_url}/reset-password/{token}"
    try:
        msg = Message(
            subject='Redefinicao de senha — Registro Original de Custos',
            recipients=[email],
            body=f"Ola {usuario['nome']},\n\nClique no link abaixo para redefinir sua senha (valido por 1 hora):\n\n{reset_url}\n\nSe nao solicitou isso, ignore este email."
        )
        mail.send(msg)
    except Exception as e:
        return jsonify({'error': f'Erro ao enviar email: {str(e)}'}), 500
    return jsonify({'ok': True})


@app.route('/reset-password/<token>')
def reset_password_page(token):
    row = get_valid_reset_token(DB_PATH, token)
    if not row:
        return render_template('reset_password.html', token=token, invalido=True)
    return render_template('reset_password.html', token=token, invalido=False)


@app.route('/api/auth/reset-password', methods=['POST'])
def api_reset_password():
    data = request.get_json(force=True) or {}
    token = data.get('token', '')
    senha_nova = data.get('senha_nova', '')
    if not token or not senha_nova:
        return jsonify({'error': 'Dados invalidos'}), 400
    if len(senha_nova) < 6:
        return jsonify({'error': 'Senha deve ter pelo menos 6 caracteres'}), 400
    row = get_valid_reset_token(DB_PATH, token)
    if not row:
        return jsonify({'error': 'Link invalido ou expirado'}), 400
    change_password(DB_PATH, row['usuario_id'], generate_password_hash(senha_nova))
    delete_reset_token(DB_PATH, token)
    return jsonify({'ok': True})


# ─── Frontend ────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html', usuario_nome=current_user.nome)


# ─── IMOVEIS ─────────────────────────────────────────────────────────────────

@app.route('/api/imoveis', methods=['GET'])
@login_required
def api_get_imoveis():
    try:
        imoveis = get_imoveis(DB_PATH, current_user.id)
        for im in imoveis:
            custos = get_custos(DB_PATH, im['id'])
            im['total_investido'] = sum(c['valor'] for c in custos)
            im['num_custos'] = len(custos)
            im['num_docs'] = sum(1 for c in custos if c.get('comprovante_path'))
        return jsonify(imoveis)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/imoveis', methods=['POST'])
@login_required
def api_add_imovel():
    try:
        data = request.get_json(force=True)
        if not data or not data.get('nome'):
            return jsonify({'error': 'Nome e obrigatorio'}), 400
        imovel = add_imovel(DB_PATH, data, current_user.id)
        send_imovel_notification(imovel['id'], 'novo_imovel')
        return jsonify(imovel), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── CUSTOS ──────────────────────────────────────────────────────────────────

@app.route('/api/custos/<imovel_id>', methods=['GET'])
@login_required
def api_get_custos(imovel_id):
    try:
        if not imovel_belongs_to_user(imovel_id, current_user.id):
            return jsonify({'error': 'Acesso negado'}), 403
        custos = get_custos(DB_PATH, imovel_id)
        return jsonify(custos)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/custos', methods=['POST'])
@login_required
def api_add_custo():
    try:
        data = {}
        if request.content_type and 'multipart' in request.content_type:
            data = request.form.to_dict()
            file = request.files.get('comprovante')
            if file and file.filename:
                unique_name, orig_name, mime = save_upload(file)
                data['comprovante_path'] = unique_name
                data['comprovante_nome'] = orig_name
                data['comprovante_tipo'] = mime
        else:
            data = request.get_json(force=True) or {}

        if not imovel_belongs_to_user(data.get('imovel_id', ''), current_user.id):
            return jsonify({'error': 'Acesso negado'}), 403

        required = ['imovel_id', 'nicho', 'valor', 'data_pagamento', 'forma_pagamento']
        for field in required:
            if not data.get(field):
                return jsonify({'error': f'Campo obrigatorio: {field}'}), 400

        custo = add_custo(DB_PATH, data)
        send_imovel_notification(custo['imovel_id'], 'novo_lancamento', custo)
        return jsonify(custo), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/custos/<custo_id>', methods=['PUT'])
@login_required
def api_update_custo(custo_id):
    try:
        if not custo_belongs_to_user(custo_id, current_user.id):
            return jsonify({'error': 'Acesso negado'}), 403

        existing = get_custo(DB_PATH, custo_id)
        if not existing:
            return jsonify({'error': 'Custo nao encontrado'}), 404

        data = {}
        new_file_uploaded = False

        if request.content_type and 'multipart' in request.content_type:
            data = request.form.to_dict()
            file = request.files.get('comprovante')
            if file and file.filename:
                old_path = existing.get('comprovante_path')
                if old_path:
                    full_old = os.path.join(UPLOAD_FOLDER, old_path)
                    if os.path.exists(full_old):
                        os.remove(full_old)
                unique_name, orig_name, mime = save_upload(file)
                data['comprovante_path'] = unique_name
                data['comprovante_nome'] = orig_name
                data['comprovante_tipo'] = mime
                new_file_uploaded = True
        else:
            data = request.get_json(force=True) or {}

        if not new_file_uploaded:
            data['comprovante_path'] = None

        custo = update_custo(DB_PATH, custo_id, data)
        return jsonify(custo)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/custos/<custo_id>', methods=['DELETE'])
@login_required
def api_delete_custo(custo_id):
    try:
        if not custo_belongs_to_user(custo_id, current_user.id):
            return jsonify({'error': 'Acesso negado'}), 403
        comp_path = delete_custo(DB_PATH, custo_id)
        if comp_path:
            full_path = os.path.join(UPLOAD_FOLDER, comp_path)
            if os.path.exists(full_path):
                os.remove(full_path)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── File serving ─────────────────────────────────────────────────────────────

@app.route('/uploads/<filename>')
@login_required
def serve_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


# ─── Excel Export ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="eef2ff", end_color="eef2ff", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11, color="0f172a")
ALT_FILL = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

thin = Side(style='thin', color='e2e8f0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = ['Data Pagamento', 'Nicho', 'Descricao', 'Favorecido', 'Forma Pagamento', 'Comprovante', 'Observacoes', 'Valor (R$)']
COL_WIDTHS = [18, 28, 38, 26, 22, 30, 36, 18]


def write_sheet(ws, imovel, custos):
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Imovel: {imovel['nome']}"
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    addr_parts = []
    if imovel.get('rua'):
        addr_parts.append(imovel['rua'])
    if imovel.get('numero'):
        addr_parts.append(f"No {imovel['numero']}")
    if imovel.get('complemento'):
        addr_parts.append(imovel['complemento'])
    if imovel.get('cep'):
        addr_parts.append(f"CEP {imovel['cep']}")
    addr_cell = ws['A2']
    addr_cell.value = ', '.join(addr_parts) if addr_parts else 'Sem endereco'
    addr_cell.font = Font(color="FFFFFF", size=10)
    addr_cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    addr_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.merge_cells('A3:H3')
    arr_cell = ws['A3']
    arr_date = imovel.get('data_arrematacao', '')
    arr_cell.value = f"Data de Arrematacao: {arr_date}" if arr_date else ''
    arr_cell.font = Font(color="FFFFFF", size=9)
    arr_cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    arr_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 14

    for col_i, (col_name, col_w) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=5, column=col_i, value=col_name)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = col_w
    ws.row_dimensions[5].height = 20

    total = 0.0
    for row_i, c in enumerate(custos, start=6):
        fill = ALT_FILL if row_i % 2 == 0 else PatternFill(fill_type=None)
        vals = [
            c.get('data_pagamento', ''),
            c.get('nicho', ''),
            c.get('descricao', ''),
            c.get('favorecido', ''),
            c.get('forma_pagamento', ''),
            c.get('comprovante_nome', ''),
            c.get('observacoes', ''),
            c.get('valor', 0.0),
        ]
        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(vertical='center', wrap_text=(col_i in [3, 7]))
            if col_i == 8:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        total += float(c.get('valor', 0))
        ws.row_dimensions[row_i].height = 16

    total_row = len(custos) + 6
    ws.merge_cells(f'A{total_row}:G{total_row}')
    label_cell = ws.cell(row=total_row, column=1, value='TOTAL INVESTIDO')
    label_cell.font = TOTAL_FONT
    label_cell.fill = TOTAL_FILL
    label_cell.alignment = Alignment(horizontal='right', vertical='center')
    label_cell.border = BORDER

    val_cell = ws.cell(row=total_row, column=8, value=total)
    val_cell.font = TOTAL_FONT
    val_cell.fill = TOTAL_FILL
    val_cell.number_format = '#,##0.00'
    val_cell.alignment = Alignment(horizontal='right', vertical='center')
    val_cell.border = BORDER
    ws.row_dimensions[total_row].height = 22

    nicho_totals = {}
    for c in custos:
        n = c.get('nicho', 'Outros')
        nicho_totals[n] = nicho_totals.get(n, 0) + float(c.get('valor', 0))

    summary_start = total_row + 2
    ws.cell(row=summary_start, column=1, value='Subtotal por Nicho').font = Font(bold=True, color='0f172a', size=10)
    ws.merge_cells(f'A{summary_start}:H{summary_start}')

    for si, (nicho, val) in enumerate(sorted(nicho_totals.items()), start=1):
        r = summary_start + si
        ws.cell(row=r, column=1, value=nicho).font = Font(bold=True)
        ws.cell(row=r, column=8, value=val).number_format = '#,##0.00'


@app.route('/api/export/excel/<imovel_id>')
@login_required
def export_excel(imovel_id):
    try:
        if not imovel_belongs_to_user(imovel_id, current_user.id):
            return jsonify({'error': 'Acesso negado'}), 403
        imoveis = get_imoveis(DB_PATH, current_user.id)
        imovel = next((i for i in imoveis if i['id'] == imovel_id), None)
        if not imovel:
            return jsonify({'error': 'Imovel nao encontrado'}), 404
        custos = get_custos(DB_PATH, imovel_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        safe_name = ''.join(c for c in imovel['nome'] if c.isalnum() or c in (' ', '-', '_'))[:28]
        ws.title = safe_name or 'Imovel'
        write_sheet(ws, imovel, custos)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"custos_{safe_name.replace(' ', '_')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel/all')
@login_required
def export_excel_all():
    try:
        imoveis = get_imoveis(DB_PATH, current_user.id)
        if not imoveis:
            return jsonify({'error': 'Nenhum imovel cadastrado'}), 404

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for imovel in imoveis:
            custos = get_custos(DB_PATH, imovel['id'])
            safe_name = ''.join(c for c in imovel['nome'] if c.isalnum() or c in (' ', '-', '_'))[:28]
            ws = wb.create_sheet(title=safe_name or f"Imovel_{imovel['id'][:6]}")
            write_sheet(ws, imovel, custos)

        ws_sum = wb.create_sheet(title='Resumo Geral', index=0)
        ws_sum.merge_cells('A1:C1')
        ws_sum['A1'].value = 'Resumo Geral — Todos os Imoveis'
        ws_sum['A1'].font = Font(color='FFFFFF', bold=True, size=14)
        ws_sum['A1'].fill = HEADER_FILL
        ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_sum.row_dimensions[1].height = 28

        headers = ['Imovel', 'Lancamentos', 'Total Investido (R$)']
        for ci, h in enumerate(headers, 1):
            cell = ws_sum.cell(row=3, column=ci, value=h)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')

        ws_sum.column_dimensions['A'].width = 40
        ws_sum.column_dimensions['B'].width = 18
        ws_sum.column_dimensions['C'].width = 24

        grand_total = 0.0
        for ri, imovel in enumerate(imoveis, start=4):
            custos = get_custos(DB_PATH, imovel['id'])
            total = sum(float(c.get('valor', 0)) for c in custos)
            grand_total += total
            fill = ALT_FILL if ri % 2 == 0 else PatternFill(fill_type=None)
            c1 = ws_sum.cell(row=ri, column=1, value=imovel['nome'])
            c1.border = BORDER
            c1.fill = fill
            c2 = ws_sum.cell(row=ri, column=2, value=len(custos))
            c2.border = BORDER
            c2.fill = fill
            c2.alignment = Alignment(horizontal='center')
            c3 = ws_sum.cell(row=ri, column=3, value=total)
            c3.border = BORDER
            c3.fill = fill
            c3.number_format = '#,##0.00'
            c3.alignment = Alignment(horizontal='right')

        grand_row = len(imoveis) + 4
        ws_sum.merge_cells(f'A{grand_row}:B{grand_row}')
        gc1 = ws_sum.cell(row=grand_row, column=1, value='TOTAL GERAL')
        gc1.font = TOTAL_FONT
        gc1.fill = TOTAL_FILL
        gc1.border = BORDER
        gc1.alignment = Alignment(horizontal='right')
        gc3 = ws_sum.cell(row=grand_row, column=3, value=grand_total)
        gc3.font = TOTAL_FONT
        gc3.fill = TOTAL_FILL
        gc3.border = BORDER
        gc3.number_format = '#,##0.00'
        gc3.alignment = Alignment(horizontal='right')

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='custos_arrematacao_todos.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── PARCEIROS ────────────────────────────────────────────────────────────────

@app.route('/api/parceiros/<imovel_id>', methods=['GET'])
@login_required
def api_get_parceiros(imovel_id):
    if not imovel_belongs_to_user(imovel_id, current_user.id):
        return jsonify({'error': 'Acesso negado'}), 403
    return jsonify(get_parceiros(DB_PATH, imovel_id))


@app.route('/api/parceiros', methods=['POST'])
@login_required
def api_add_parceiro():
    data = request.get_json(force=True) or {}
    imovel_id = data.get('imovel_id', '')
    email = (data.get('email') or '').strip().lower()
    nome = (data.get('nome') or '').strip()
    notificar = data.get('notificar', True)
    if not imovel_id or not email:
        return jsonify({'error': 'imovel_id e email sao obrigatorios'}), 400
    if not imovel_belongs_to_user(imovel_id, current_user.id):
        return jsonify({'error': 'Acesso negado'}), 403
    parceiro = add_parceiro(DB_PATH, imovel_id, nome, email, notificar)
    return jsonify(parceiro), 201


@app.route('/api/parceiros/<parceiro_id>', methods=['DELETE'])
@login_required
def api_delete_parceiro(parceiro_id):
    delete_parceiro(DB_PATH, parceiro_id)
    return jsonify({'ok': True})


@app.route('/api/parceiros/<parceiro_id>', methods=['PATCH'])
@login_required
def api_toggle_parceiro(parceiro_id):
    data = request.get_json(force=True) or {}
    update_parceiro_notificar(DB_PATH, parceiro_id, data.get('notificar', True))
    return jsonify({'ok': True})


# ─── EMAIL NOTIFICATIONS ───────────────────────────────────────────────────────

def _build_email_html(imovel, custos, event_type, new_custo=None):
    addr_parts = []
    if imovel.get('rua'): addr_parts.append(imovel['rua'])
    if imovel.get('numero'): addr_parts.append(f"No {imovel['numero']}")
    if imovel.get('cep'): addr_parts.append(f"CEP {imovel['cep']}")
    addr = ', '.join(addr_parts) if addr_parts else ''

    total = sum(float(c.get('valor', 0)) for c in custos)

    nicho_totals = {}
    for c in custos:
        n = c.get('nicho', 'Outros')
        nicho_totals[n] = nicho_totals.get(n, 0) + float(c.get('valor', 0))

    def fmt(v):
        return f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    if event_type == 'novo_imovel':
        event_color = '#2563eb'
        event_label = '🏠 Novo Imóvel Cadastrado'
        event_desc = 'Um novo imóvel foi adicionado ao sistema.'
    else:
        event_color = '#10b981'
        event_label = '➕ Novo Lançamento Registrado'
        event_desc = 'Um novo custo foi lançado neste imóvel.'

    new_item_html = ''
    if new_custo:
        new_item_html = f"""
        <tr><td style="background:#f0fdf4;padding:20px 32px;border-bottom:1px solid #e2e8f0">
          <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:#15803d;margin-bottom:10px">Novo Lançamento</div>
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:10px 14px;background:#fff;border-radius:8px;border:1px solid #bbf7d0">
                <table width="100%" cellpadding="4" cellspacing="0">
                  <tr><td style="font-size:12px;color:#64748b;width:120px">Nicho</td><td style="font-size:13px;font-weight:600">{new_custo.get('nicho','')}</td></tr>
                  <tr><td style="font-size:12px;color:#64748b">Descrição</td><td style="font-size:13px;font-weight:600">{new_custo.get('descricao','')}</td></tr>
                  <tr><td style="font-size:12px;color:#64748b">Valor</td><td style="font-size:15px;font-weight:800;color:#15803d">{fmt(float(new_custo.get('valor',0)))}</td></tr>
                  <tr><td style="font-size:12px;color:#64748b">Data</td><td style="font-size:13px">{new_custo.get('data_pagamento','')}</td></tr>
                  <tr><td style="font-size:12px;color:#64748b">Pagamento</td><td style="font-size:13px">{new_custo.get('forma_pagamento','')}</td></tr>
                  {f'<tr><td style="font-size:12px;color:#64748b">Favorecido</td><td style="font-size:13px">{new_custo.get("favorecido","")}</td></tr>' if new_custo.get('favorecido') else ''}
                </table>
              </td>
            </tr>
          </table>
        </td></tr>"""

    nicho_rows = ''.join(
        f'<tr><td style="padding:7px 0;font-size:13px;border-bottom:1px solid #f1f5f9">{n}</td>'
        f'<td style="padding:7px 0;font-size:13px;font-weight:700;text-align:right;border-bottom:1px solid #f1f5f9">{fmt(v)}</td></tr>'
        for n, v in sorted(nicho_totals.items(), key=lambda x: -x[1])
    ) if nicho_totals else '<tr><td colspan="2" style="font-size:13px;color:#64748b;padding:8px 0">Sem lançamentos ainda.</td></tr>'

    arr = imovel.get('data_arrematacao', '')
    arr_cell = f'<td width="12"></td><td align="center" style="padding:12px;background:#f8fafc;border-radius:10px;border:1px solid #e2e8f0"><div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.5px">Arrematação</div><div style="font-size:15px;font-weight:800;color:#0f172a;margin-top:4px">{arr}</div></td>' if arr else ''

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"/></head>
<body style="margin:0;padding:0;background:#f1f5f9;font-family:Segoe UI,Helvetica,Arial,sans-serif;color:#0f172a">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:24px 16px">
<tr><td align="center"><table width="580" cellpadding="0" cellspacing="0" style="width:100%;max-width:580px">

<tr><td style="background:#1e3a8a;border-radius:14px 14px 0 0;padding:28px 32px">
  <div style="color:#a5b4fc;font-size:10px;text-transform:uppercase;letter-spacing:1.2px;font-weight:700">LeilaoPro · Notificação</div>
  <div style="color:#fff;font-size:22px;font-weight:800;margin-top:8px;line-height:1.2">{imovel.get('nome','')}</div>
  {f'<div style="color:rgba(255,255,255,.6);font-size:12px;margin-top:6px">{addr}</div>' if addr else ''}
</td></tr>

<tr><td style="background:{event_color};padding:13px 32px">
  <span style="color:#fff;font-size:13px;font-weight:700">{event_label}</span>
  <span style="color:rgba(255,255,255,.75);font-size:12px;margin-left:10px">{event_desc}</span>
</td></tr>

<tr><td style="background:#fff;padding:24px 32px;border-bottom:1px solid #e2e8f0">
  <table width="100%" cellpadding="0" cellspacing="0"><tr>
    <td align="center" style="padding:14px;background:#f8fafc;border-radius:10px;border:1px solid #e2e8f0">
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.5px">Total Investido</div>
      <div style="font-size:22px;font-weight:800;color:#0f172a;margin-top:4px">{fmt(total)}</div>
    </td>
    <td width="12"></td>
    <td align="center" style="padding:14px;background:#f8fafc;border-radius:10px;border:1px solid #e2e8f0">
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.5px">Lançamentos</div>
      <div style="font-size:22px;font-weight:800;color:#0f172a;margin-top:4px">{len(custos)}</div>
    </td>
    {arr_cell}
  </tr></table>
</td></tr>

{new_item_html}

<tr><td style="background:#fff;padding:20px 32px 28px">
  <div style="font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:#64748b;margin-bottom:12px">Breakdown por Nicho</div>
  <table width="100%" cellpadding="0" cellspacing="0">{nicho_rows}</table>
</td></tr>

<tr><td style="background:#1e293b;border-radius:0 0 14px 14px;padding:18px 32px;text-align:center">
  <div style="color:rgba(255,255,255,.55);font-size:11px">LeilaoPro · Registro Original de Custos</div>
  <div style="color:rgba(255,255,255,.3);font-size:10px;margin-top:4px">Você está recebendo este e-mail pois é parceiro neste imóvel.</div>
</td></tr>

</table></td></tr></table>
</body></html>"""


def send_imovel_notification(imovel_id, event_type, new_custo=None):
    if not app.config.get('MAIL_USERNAME'):
        return

    def _send():
        with app.app_context():
            try:
                imovel = get_imovel_by_id(DB_PATH, imovel_id)
                if not imovel:
                    return
                custos = get_custos(DB_PATH, imovel_id)
                parceiros = [p for p in get_parceiros(DB_PATH, imovel_id) if p.get('notificar')]
                if not parceiros:
                    return
                subject = f"[LeilaoPro] {'Novo imóvel' if event_type == 'novo_imovel' else 'Novo lançamento'} — {imovel.get('nome','')}"
                html_body = _build_email_html(imovel, custos, event_type, new_custo)
                for p in parceiros:
                    try:
                        msg = Message(subject=subject, recipients=[p['email']], html=html_body)
                        mail.send(msg)
                    except Exception:
                        pass
            except Exception:
                pass

    threading.Thread(target=_send, daemon=True).start()


# ─── VIABILIDADE FINANCEIRA ───────────────────────────────────────────────────

VIABILIDADE_PROMPT = """Você é um especialista em leilões imobiliários no Brasil, com experiência em leilões extrajudiciais de bancos, leilões judiciais, análise de matrícula imobiliária, análise de edital de leilão e investimento imobiliário para revenda (flip).

O objetivo da análise é identificar se o imóvel representa uma boa oportunidade de investimento para revenda rápida, buscando ROI mínimo de 30%.

Analise o edital e a matrícula como um investidor profissional, procurando riscos jurídicos ocultos, custos reais da operação e potencial real de lucro.

REGRAS DE CUSTOS (OBRIGATÓRIO):
- R$ 30.000 para obra (sempre incluir)
- R$ 15.000 para desocupação/imissão na posse (incluir se ocupado ou situação não clara; desconsiderar se edital confirmar imóvel desocupado)
- Estimar: ITBI, registro em cartório, taxas cartorárias, certidões, comissão do leiloeiro

ANÁLISE OBRIGATÓRIA DA MATRÍCULA — leia linha por linha e identifique: alienação fiduciária, consolidação da propriedade, hipoteca, penhoras, usufruto, indisponibilidade, execuções fiscais, averbações, ações judiciais, risco de ocupação, disputa judicial, nulidade do leilão.

RESPONSABILIDADE POR DÍVIDAS — identifique quem paga: IPTU, condomínio, débitos fiscais, taxas municipais, dívidas anteriores.

ANÁLISE DE MERCADO (REGRA ABSOLUTA) — NÃO utilize o valor de avaliação do banco/leiloeiro. Use exclusivamente pesquisa de imóveis comparáveis (OLX, Viva Real, ZAP Imóveis, Imovelweb) considerando mesma região, metragem, tipo, padrão, quartos e vagas. Se não encontrar comparáveis suficientes, pergunte ao usuário o valor projetado de venda.

Apresente a resposta exatamente nesta estrutura:

1️⃣ DADOS DO IMÓVEL
2️⃣ ANÁLISE DE RISCO
3️⃣ CUSTOS TOTAIS ESTIMADOS
4️⃣ SIMULAÇÃO DE REVENDA COM ROI (cenários conservador, provável e otimista)
5️⃣ ANÁLISE ESTRATÉGICA (liquidez, perfil comprador, tempo de venda)
6️⃣ SCORE DE OPORTUNIDADE (0-100): 80-100 Excelente | 60-79 Boa | 40-59 Moderada | 0-39 Alto risco
7️⃣ LANCE MÁXIMO RECOMENDADO (para ROI mínimo de 30%)
8️⃣ VALIDAÇÃO DE INFORMAÇÕES CRÍTICAS
9️⃣ CONCLUSÃO FINAL

Ao final, apresente um bloco estruturado:
CIDADE:
BAIRRO:
TIPO_IMOVEL:
METRAGEM_APROXIMADA:
VALOR_ARREMATACAO_ESTIMADO:
CUSTO_TOTAL_OPERACAO:
VALOR_MERCADO_ESTIMADO:
VALOR_VENDA_CONSERVADOR:
VALOR_VENDA_PROVAVEL:
VALOR_VENDA_OTIMISTA:
LUCRO_LIQUIDO_ESTIMADO:
ROI_PROJETADO:
SCORE_OPORTUNIDADE:
SITUACAO_OCUPACAO:
RISCO_JURIDICO_PRINCIPAL:
LIQUIDEZ_REGIAO:

Quando uma informação não puder ser confirmada pelos documentos, escreva explicitamente: "Esta informação não pôde ser confirmada apenas com os documentos fornecidos." Nunca interrompa a análise por isso — continue com as informações disponíveis."""


@app.route('/api/viabilidade/analisar', methods=['POST'])
@login_required
def api_viabilidade():
    import base64
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return jsonify({'error': 'API de análise não configurada. Adicione ANTHROPIC_API_KEY nas variáveis de ambiente.'}), 503

    edital = request.files.get('edital')
    matricula = request.files.get('matricula')
    if not edital:
        return jsonify({'error': 'Edital do leilão é obrigatório (PDF)'}), 400
    if not matricula:
        return jsonify({'error': 'Matrícula do imóvel é obrigatória (PDF)'}), 400

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        edital_b64 = base64.standard_b64encode(edital.read()).decode('utf-8')
        matricula_b64 = base64.standard_b64encode(matricula.read()).decode('utf-8')

        message = client.messages.create(
            model='claude-opus-4-6',
            max_tokens=8000,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'text', 'text': 'Documento 1 — Edital do Leilão (PDF):'},
                    {'type': 'document', 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': edital_b64}},
                    {'type': 'text', 'text': 'Documento 2 — Matrícula do Imóvel (PDF):'},
                    {'type': 'document', 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': matricula_b64}},
                    {'type': 'text', 'text': VIABILIDADE_PROMPT}
                ]
            }]
        )
        analise_texto = message.content[0].text
        saved = save_viabilidade(
            DB_PATH, current_user.id, analise_texto,
            nome_edital=edital.filename, nome_matricula=matricula.filename
        )
        return jsonify({'ok': True, 'analise': analise_texto, 'analise_id': saved['id']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/viabilidade/historico', methods=['GET'])
@login_required
def api_viabilidade_historico():
    rows = get_viabilidades(DB_PATH, current_user.id)
    return jsonify(rows)


@app.route('/api/viabilidade/<analise_id>', methods=['GET'])
@login_required
def api_viabilidade_get(analise_id):
    row = get_viabilidade_by_id(DB_PATH, analise_id, current_user.id)
    if not row:
        return jsonify({'error': 'Não encontrado'}), 404
    return jsonify(row)


@app.route('/api/viabilidade/<analise_id>', methods=['DELETE'])
@login_required
def api_viabilidade_delete(analise_id):
    delete_viabilidade(DB_PATH, analise_id, current_user.id)
    return jsonify({'ok': True})


@app.route('/api/viabilidade/chat', methods=['POST'])
@login_required
def api_viabilidade_chat():
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return jsonify({'error': 'API não configurada'}), 503
    data = request.get_json(force=True) or {}
    historico = data.get('historico', [])
    analise_original = data.get('analise_original', '')
    if not historico:
        return jsonify({'error': 'Mensagem obrigatória'}), 400
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        system_msg = f"""Você é o assistente de análise de leilões imobiliários do LeilaoPro.
O usuário já recebeu uma análise completa de viabilidade financeira.
Quando o usuário informar novos dados (área corrigida, valor de mercado, situação de ocupação, etc.),
recalcule os valores afetados e apresente os resultados atualizados de forma clara e objetiva.
Mantenha o mesmo rigor técnico da análise original.

ANÁLISE ORIGINAL:
{analise_original[:3000]}"""
        message = client.messages.create(
            model='claude-opus-4-6',
            max_tokens=4000,
            system=system_msg,
            messages=historico
        )
        return jsonify({'ok': True, 'resposta': message.content[0].text})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/terms')
def terms_page():
    return render_template('terms.html')


@app.route('/privacy')
def privacy_page():
    return render_template('privacy.html')


if __name__ == '__main__':
    init_db(DB_PATH)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    port = int(os.environ.get('PORT', 5000))
    is_local = port == 5000
    if is_local:
        import webbrowser, threading
        threading.Timer(1.2, lambda: webbrowser.open('http://localhost:5000')).start()
    app.run(debug=False, host='0.0.0.0', port=port)
